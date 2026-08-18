from io import BytesIO

from django.urls import reverse
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from apps.authentication.models import User
from apps.companies.models import ClientProfile, Company
from apps.content_calendar.excel import TEMPLATE_HEADERS
from apps.content_calendar.models import ContentCalendarItem


def _build_workbook(rows, headers=None):
    wb = Workbook()
    ws = wb.active
    ws.append(headers or TEMPLATE_HEADERS)
    for row in rows:
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    buffer.name = 'calendar.xlsx'
    return buffer


# Date | Day | Platform | Format | Topical | Weekly Theme | Post Time |
# Caption / Content Idea | Visual Brief | Hashtags + CTA | Status | Source
VALID_ROW = [
    '2026-10-18', 'Sun', 'Instagram, Facebook, X, Pinterest, LinkedIn', 'Complimentary Creative',
    'Diwali sale post', 'Festive Sales', '10:00',
    'Mention 20% off', 'Festive product shot',
    '#diwali #sale | CTA: Shop now', 'Posted', '',
]

INVALID_ROW = [
    'not-a-date', 'Wed', '', '',
    '', '', '',
    '', '', '', 'NotAStatus', '',
]


class BaseContentCalendarTestCase(APITestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(email='admin@example.com', password='StrongPass123!')
        self.company = Company.objects.create(name='Acme Retail', created_by=self.admin)
        self.other_company = Company.objects.create(name='Other Co', created_by=self.admin)

        self.client_user = User.objects.create_user(
            email='client@example.com', password='StrongPass123!', role=User.Role.CLIENT,
        )
        ClientProfile.objects.create(user=self.client_user, company=self.company)

        self.item = ContentCalendarItem.objects.create(
            company=self.company,
            topic='Existing post',
            content_type='Static Post',
            platforms=['instagram'],
            scheduled_date='2026-09-01',
            created_by=self.admin,
        )

    def authenticate_as(self, email, password):
        response = self.client.post(reverse('authentication:login'), {'email': email, 'password': password})
        access = response.data['access']
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access}')
        return response

    def list_url(self, company_id=None):
        return reverse('content_calendar:item-list-create', kwargs={'company_id': company_id or self.company.pk})

    def detail_url(self, pk, company_id=None):
        return reverse('content_calendar:item-detail', kwargs={'company_id': company_id or self.company.pk, 'pk': pk})


class CalendarItemCrudTests(BaseContentCalendarTestCase):
    def test_admin_can_create_item_with_freeform_category_and_format(self):
        self.authenticate_as('admin@example.com', 'StrongPass123!')
        response = self.client.post(self.list_url(), {
            'topic': 'New Year post',
            'category': 'Whatever label the client uses',
            'weekly_theme': 'Leadership',
            'content_type': 'Static + Story',
            'platforms': ['instagram', 'linkedin', 'twitter', 'pinterest'],
            'scheduled_date': '2026-12-31',
            'hashtags': ['#newyear'],
        }, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        item = ContentCalendarItem.objects.get(topic='New Year post')
        self.assertEqual(item.company, self.company)
        self.assertEqual(item.source, ContentCalendarItem.Source.MANUAL)
        self.assertEqual(item.status, ContentCalendarItem.Status.DRAFT)
        self.assertEqual(item.category, 'Whatever label the client uses')
        self.assertEqual(item.weekly_theme, 'Leadership')
        self.assertIn('twitter', item.platforms)
        self.assertIn('pinterest', item.platforms)

    def test_category_and_weekly_theme_are_optional(self):
        self.authenticate_as('admin@example.com', 'StrongPass123!')
        response = self.client.post(self.list_url(), {
            'topic': 'No theme post', 'content_type': 'Reel',
            'platforms': ['instagram'], 'scheduled_date': '2026-12-31',
        }, format='json')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)

    def test_create_requires_valid_platform(self):
        self.authenticate_as('admin@example.com', 'StrongPass123!')
        response = self.client.post(self.list_url(), {
            'topic': 'Bad platform post', 'content_type': 'Reel',
            'platforms': ['mastodon'], 'scheduled_date': '2026-12-31',
        }, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_client_cannot_create_item(self):
        self.authenticate_as('client@example.com', 'StrongPass123!')
        response = self.client.post(self.list_url(), {
            'topic': 'Should fail', 'content_type': 'Reel',
            'platforms': ['instagram'], 'scheduled_date': '2026-12-31',
        }, format='json')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_client_can_view_own_company_calendar(self):
        self.authenticate_as('client@example.com', 'StrongPass123!')
        response = self.client.get(self.list_url())
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['count'], 1)

    def test_client_cannot_view_other_company_calendar(self):
        self.authenticate_as('client@example.com', 'StrongPass123!')
        response = self.client.get(self.list_url(company_id=self.other_company.pk))
        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_admin_can_edit_item(self):
        self.authenticate_as('admin@example.com', 'StrongPass123!')
        response = self.client.patch(self.detail_url(self.item.pk), {'status': 'scheduled'})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.item.refresh_from_db()
        self.assertEqual(self.item.status, ContentCalendarItem.Status.SCHEDULED)

    def test_admin_can_delete_item(self):
        self.authenticate_as('admin@example.com', 'StrongPass123!')
        response = self.client.delete(self.detail_url(self.item.pk))
        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)
        self.assertFalse(ContentCalendarItem.objects.filter(pk=self.item.pk).exists())

    def test_filter_by_status_and_platform(self):
        ContentCalendarItem.objects.create(
            company=self.company, topic='Scheduled one', content_type='Story',
            platforms=['linkedin'], scheduled_date='2026-09-05', status=ContentCalendarItem.Status.SCHEDULED,
        )
        self.authenticate_as('admin@example.com', 'StrongPass123!')

        response = self.client.get(self.list_url(), {'status': 'scheduled'})
        self.assertEqual(response.data['count'], 1)

        response = self.client.get(self.list_url(), {'platform': 'linkedin'})
        self.assertEqual(response.data['count'], 1)

        response = self.client.get(self.list_url(), {'platform': 'instagram'})
        self.assertEqual(response.data['count'], 1)

    def test_filter_by_month(self):
        # self.item is scheduled for 2026-09-01.
        ContentCalendarItem.objects.create(
            company=self.company, topic='October item', content_type='Reel',
            platforms=['instagram'], scheduled_date='2026-10-15',
        )
        self.authenticate_as('admin@example.com', 'StrongPass123!')

        response = self.client.get(self.list_url(), {'month': '2026-09'})
        self.assertEqual(response.data['count'], 1)
        self.assertEqual(response.data['results'][0]['topic'], 'Existing post')

        response = self.client.get(self.list_url(), {'month': '2026-10'})
        self.assertEqual(response.data['count'], 1)
        self.assertEqual(response.data['results'][0]['topic'], 'October item')

        response = self.client.get(self.list_url(), {'month': '2026-11'})
        self.assertEqual(response.data['count'], 0)


class DuplicateTests(BaseContentCalendarTestCase):
    def test_admin_can_duplicate_item(self):
        self.authenticate_as('admin@example.com', 'StrongPass123!')
        url = reverse('content_calendar:item-duplicate', kwargs={'company_id': self.company.pk, 'pk': self.item.pk})
        response = self.client.post(url)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(ContentCalendarItem.objects.filter(company=self.company).count(), 2)


class TemplateDownloadTests(BaseContentCalendarTestCase):
    def test_admin_can_download_template(self):
        self.authenticate_as('admin@example.com', 'StrongPass123!')
        url = reverse('content_calendar:template', kwargs={'company_id': self.company.pk})
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('spreadsheetml', response['Content-Type'])
        wb = load_workbook(BytesIO(response.content))
        header_row = [cell.value for cell in next(wb.active.iter_rows(max_row=1))]
        self.assertEqual(header_row, TEMPLATE_HEADERS)

    def test_client_cannot_download_template(self):
        self.authenticate_as('client@example.com', 'StrongPass123!')
        url = reverse('content_calendar:template', kwargs={'company_id': self.company.pk})
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)


class ImportTests(BaseContentCalendarTestCase):
    def test_preview_reports_valid_and_invalid_rows_without_saving(self):
        self.authenticate_as('admin@example.com', 'StrongPass123!')
        workbook = _build_workbook([VALID_ROW, INVALID_ROW])
        url = reverse('content_calendar:import-preview', kwargs={'company_id': self.company.pk})
        response = self.client.post(url, {'file': workbook}, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['valid_count'], 1)
        self.assertEqual(response.data['invalid_count'], 1)
        self.assertTrue(response.data['invalid_rows'][0]['errors'])
        # Preview must not touch the database.
        self.assertEqual(ContentCalendarItem.objects.filter(company=self.company).count(), 1)

    def test_commit_creates_valid_rows_with_parsed_fields(self):
        self.authenticate_as('admin@example.com', 'StrongPass123!')
        workbook = _build_workbook([VALID_ROW, INVALID_ROW])
        url = reverse('content_calendar:import-commit', kwargs={'company_id': self.company.pk})
        response = self.client.post(url, {'file': workbook}, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['created_count'], 1)
        self.assertEqual(response.data['invalid_count'], 1)

        created = ContentCalendarItem.objects.get(topic='Diwali sale post')
        self.assertEqual(created.source, ContentCalendarItem.Source.EXCEL_IMPORT)
        self.assertEqual(set(created.platforms), {'instagram', 'facebook', 'twitter', 'pinterest', 'linkedin'})
        self.assertEqual(created.hashtags, ['#diwali', '#sale'])
        self.assertEqual(created.cta, 'Shop now')
        self.assertEqual(created.status, ContentCalendarItem.Status.PUBLISHED)
        self.assertEqual(created.weekly_theme, 'Festive Sales')
        self.assertEqual(created.content_type, 'Complimentary Creative')

    def test_import_handles_day_only_date_by_assuming_current_year(self):
        self.authenticate_as('admin@example.com', 'StrongPass123!')
        row = [
            '2-Aug', 'Sun', 'LinkedIn', 'Static Post', 'Short date row', '', '09:00',
            '', '', '#test', 'Draft', '',
        ]
        workbook = _build_workbook([row])
        url = reverse('content_calendar:import-commit', kwargs={'company_id': self.company.pk})
        response = self.client.post(url, {'file': workbook}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['created_count'], 1)
        created = ContentCalendarItem.objects.get(topic='Short date row')
        self.assertEqual(created.scheduled_date.month, 8)
        self.assertEqual(created.scheduled_date.day, 2)

    def test_import_handles_different_column_order_and_ampersand_platforms(self):
        # A second real-world sheet: no Status/Source columns, Topical moved to
        # the end, and platforms joined with "&" instead of ",".
        alt_headers = [
            'Date', 'Day', 'Platform', 'Format', 'Weekly Theme', 'Post Time',
            'Caption / Content Idea', 'Visual Brief', 'Hashtags + CTA', 'Topical',
        ]
        alt_row = [
            '2026-07-01', 'Wed', 'Instagram & Facebook', 'Static + Story', 'Gratitude & Service', '10:00',
            'Celebrating the hands that heal.', 'Elegant white and gold creative.',
            '#NationalDoctorsDay #ThankYouDoctors', 'National Doctor\'s Day',
        ]
        self.authenticate_as('admin@example.com', 'StrongPass123!')
        workbook = _build_workbook([alt_row], headers=alt_headers)
        url = reverse('content_calendar:import-commit', kwargs={'company_id': self.company.pk})
        response = self.client.post(url, {'file': workbook}, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['created_count'], 1)
        created = ContentCalendarItem.objects.get(topic="National Doctor's Day")
        self.assertEqual(set(created.platforms), {'instagram', 'facebook'})
        self.assertEqual(created.content_type, 'Static + Story')
        self.assertEqual(created.weekly_theme, 'Gratitude & Service')
        self.assertEqual(created.caption_requirements, 'Celebrating the hands that heal.')
        # No Status column in this sheet -> defaults to draft rather than erroring.
        self.assertEqual(created.status, ContentCalendarItem.Status.DRAFT)

    def test_import_requires_file(self):
        self.authenticate_as('admin@example.com', 'StrongPass123!')
        url = reverse('content_calendar:import-preview', kwargs={'company_id': self.company.pk})
        response = self.client.post(url, {}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_client_cannot_import(self):
        self.authenticate_as('client@example.com', 'StrongPass123!')
        workbook = _build_workbook([VALID_ROW])
        url = reverse('content_calendar:import-commit', kwargs={'company_id': self.company.pk})
        response = self.client.post(url, {'file': workbook}, format='multipart')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
