"""Excel template generation and import parsing for the content calendar (Epic 04).

The column layout mirrors the planning sheets the team already produces in
practice (Date / Day / Platform / Format / Topical / Weekly Theme / Post Time /
Caption-Content Idea / Visual Brief / Hashtags+CTA / Status / Source), so an
existing sheet can be uploaded with little to no reformatting.
"""

import datetime
import re
from io import BytesIO

from openpyxl import Workbook, load_workbook

from .models import ContentCalendarItem

TEMPLATE_HEADERS = [
    'Date', 'Day', 'Platform', 'Format', 'Topical', 'Weekly Theme', 'Post Time',
    'Caption / Content Idea', 'Visual Brief', 'Hashtags + CTA', 'Status', 'Source',
]

EXAMPLE_ROWS = [
    [
        '2026-08-02', 'Sun', 'Instagram, Facebook, X, Pinterest, LinkedIn', 'Complimentary Creative',
        'Director Birthday - Mr. Sanjay Mehta', 'Leadership', '09:00',
        'Celebrating the vision behind every landmark. Wishing Mr. Sanjay Mehta, Director, Majestique '
        'Landmarks, a very Happy Birthday. May the year ahead bring continued success, inspiring '
        'milestones and many more landmarks to celebrate.',
        'Premium leadership portrait of Mr. Sanjay Mehta. Minimal editorial composition using Majestique '
        'brand colours. Sophisticated and corporate rather than overly festive.',
        '#HappyBirthday #SanjayMehta #MajestiqueLandmarks #Leadership', 'Posted', '',
    ],
    [
        '2026-08-02', 'Sun', 'Instagram, Facebook, X, Pinterest, LinkedIn', 'Complimentary Creative',
        'International Friendship Day', 'Community', '18:00',
        'Some neighbours become friends. Some friends become family. Celebrating the connections that '
        'make a community feel like home.',
        'Warm, candid community moment with neighbours/friends. Premium and minimal rather than overly '
        'festive.',
        '#FriendshipDay #CommunityLiving #LiveMajestique | CTA: Celebrating the people who make every '
        'place feel like home.', 'Draft', '',
    ],
]

# Status aliases seen in real planning sheets, mapped onto our workflow states.
STATUS_ALIASES = {
    'posted': ContentCalendarItem.Status.PUBLISHED,
    'published': ContentCalendarItem.Status.PUBLISHED,
    'live': ContentCalendarItem.Status.PUBLISHED,
    'scheduled': ContentCalendarItem.Status.SCHEDULED,
    'approved': ContentCalendarItem.Status.APPROVED,
    'rejected': ContentCalendarItem.Status.REJECTED,
    'failed': ContentCalendarItem.Status.FAILED,
    'draft': ContentCalendarItem.Status.DRAFT,
    'pending approval': ContentCalendarItem.Status.PENDING_APPROVAL,
    'generated': ContentCalendarItem.Status.GENERATED,
    'generating': ContentCalendarItem.Status.GENERATING,
}

PLATFORM_LOOKUP = {}
for _value, _label in ContentCalendarItem.Platform.choices:
    PLATFORM_LOOKUP[_value.lower()] = _value
    PLATFORM_LOOKUP[_label.lower()] = _value
PLATFORM_LOOKUP['x'] = ContentCalendarItem.Platform.TWITTER
PLATFORM_LOOKUP['twitter'] = ContentCalendarItem.Platform.TWITTER

HASHTAG_RE = re.compile(r'#\w+')
PLATFORM_SPLIT_RE = re.compile(r'[,&/]')

# Different sheets order and word their columns differently (one file put
# "Topical" right after Format, another put it last; separators vary between
# "Hashtags + CTA" and "Hashtags & CTA"). Match by header *name*, not
# position, so column order in the uploaded file never matters.
FIELD_ALIASES = {
    'date': ['date'],
    'platform': ['platform', 'platforms'],
    'content_type': ['format'],
    'topic': ['topical', 'topic'],
    'weekly_theme': ['weekly theme', 'theme'],
    'time': ['post time', 'time'],
    'caption': ['caption content idea', 'caption', 'content idea'],
    'visual_brief': ['visual brief', 'creative brief', 'creative requirements'],
    'hashtags_cta': ['hashtags cta', 'hashtags'],
    'status': ['status'],
    'source': ['source'],
}


def _normalize_header(text):
    text = str(text or '').strip().lower()
    text = re.sub(r'[/&+]', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def _map_columns(header_row):
    """Maps canonical field names to the column index that matches them in this sheet."""
    normalized = [_normalize_header(cell) for cell in header_row]
    column_map = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                column_map[field] = normalized.index(alias)
                break
    return column_map


def build_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Content Calendar'
    ws.append(TEMPLATE_HEADERS)
    for row in EXAMPLE_ROWS:
        ws.append(row)

    for column_cells in ws.columns:
        longest = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(longest + 2, 12), 45)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _parse_date(value):
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str) and value.strip():
        text = value.strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%d-%b-%Y', '%d %b %Y', '%d %B %Y'):
            try:
                return datetime.datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        # Sheets often omit the year for a "this month/this week" plan (e.g. "2-Aug").
        for fmt in ('%d-%b', '%d %b', '%d %B'):
            try:
                parsed = datetime.datetime.strptime(text, fmt)
                return parsed.replace(year=datetime.date.today().year).date()
            except ValueError:
                continue
    return None


def _parse_time(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime.time):
        return value
    if isinstance(value, datetime.datetime):
        return value.time()
    if isinstance(value, str) and value.strip():
        for fmt in ('%H:%M', '%H:%M:%S', '%I:%M %p'):
            try:
                return datetime.datetime.strptime(value.strip(), fmt).time()
            except ValueError:
                continue
    return None


def _clean_text(value):
    return str(value).strip() if value not in (None, '') else ''


def _parse_platforms(raw):
    platforms, errors = [], []
    if raw:
        for part in PLATFORM_SPLIT_RE.split(str(raw)):
            part = part.strip().lower()
            if not part:
                continue
            platform = PLATFORM_LOOKUP.get(part)
            if platform:
                if platform not in platforms:
                    platforms.append(platform)
            else:
                errors.append(f'Platform "{part}" is not recognized.')
    return platforms, errors


def _parse_hashtags_and_cta(raw):
    """Splits a combined "Hashtags + CTA" cell, e.g. '#a #b | CTA: Shop now'."""
    if not raw:
        return [], ''
    text = str(raw)
    cta = ''
    match = re.search(r'\|?\s*CTA\s*:\s*(.+)$', text, flags=re.IGNORECASE)
    if match:
        cta = match.group(1).strip()
        text = text[:match.start()]
    hashtags = HASHTAG_RE.findall(text)
    return hashtags, cta


def _parse_status(raw):
    if not raw:
        return ContentCalendarItem.Status.DRAFT, None
    key = str(raw).strip().lower()
    if key in STATUS_ALIASES:
        return STATUS_ALIASES[key], None
    valid_values = {value.lower() for value, _label in ContentCalendarItem.Status.choices}
    if key in valid_values:
        return key, None
    return None, f'Status "{raw}" is not recognized.'


class WorkbookParseError(Exception):
    pass


def parse_and_validate(file):
    """Parse an uploaded .xlsx file and validate each row.

    Returns (valid_rows, invalid_rows), where each entry is a dict with a
    'row' (1-indexed spreadsheet row number) and 'data' (cleaned field
    values). Invalid entries additionally carry an 'errors' list.
    """
    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as exc:
        raise WorkbookParseError('Could not read this file. Make sure it is a valid .xlsx file.') from exc

    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise WorkbookParseError('This file has no header row.')

    columns = _map_columns(header_row)

    def cell(row, field):
        idx = columns.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    valid_rows = []
    invalid_rows = []

    for index, row in enumerate(rows_iter, start=2):
        if row is None or all(c in (None, '') for c in row):
            continue

        date_raw = cell(row, 'date')
        platform_raw = cell(row, 'platform')
        format_raw = cell(row, 'content_type')
        topical_raw = cell(row, 'topic')
        weekly_theme_raw = cell(row, 'weekly_theme')
        time_raw = cell(row, 'time')
        caption_raw = cell(row, 'caption')
        visual_brief_raw = cell(row, 'visual_brief')
        hashtags_cta_raw = cell(row, 'hashtags_cta')
        status_raw = cell(row, 'status')
        source_raw = cell(row, 'source')

        errors = []

        topic = _clean_text(topical_raw)
        if not topic:
            errors.append('Topical is required.')

        content_type = _clean_text(format_raw)
        if not content_type:
            errors.append('Format is required.')

        platforms, platform_errors = _parse_platforms(platform_raw)
        errors.extend(platform_errors)
        if not platforms:
            errors.append('At least one recognized platform is required.')

        scheduled_date = _parse_date(date_raw)
        if not scheduled_date:
            errors.append(f'Date "{date_raw}" could not be parsed. Use YYYY-MM-DD or "2-Aug".')

        scheduled_time = _parse_time(time_raw)
        if time_raw not in (None, '') and not scheduled_time:
            errors.append(f'Post Time "{time_raw}" could not be parsed. Use HH:MM.')

        hashtags, cta = _parse_hashtags_and_cta(hashtags_cta_raw)

        status, status_error = _parse_status(status_raw)
        if status_error:
            errors.append(status_error)

        row_data = {
            'topic': topic,
            'content_type': content_type,
            'weekly_theme': _clean_text(weekly_theme_raw),
            'platforms': platforms,
            'scheduled_date': scheduled_date,
            'scheduled_time': scheduled_time,
            'caption_requirements': _clean_text(caption_raw),
            'creative_requirements': _clean_text(visual_brief_raw),
            'cta': cta,
            'hashtags': hashtags,
            'source_notes': _clean_text(source_raw),
            'status': status or ContentCalendarItem.Status.DRAFT,
        }

        if errors:
            invalid_rows.append({'row': index, 'errors': errors, 'data': row_data})
        else:
            valid_rows.append({'row': index, 'data': row_data})

    return valid_rows, invalid_rows
